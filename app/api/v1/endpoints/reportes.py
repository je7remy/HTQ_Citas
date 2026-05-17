"""Generación de reportes en PDF con WeasyPrint y Excel con openpyxl."""
from datetime import date as date_type
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from jinja2 import Template
from sqlmodel import Session, select
from weasyprint import HTML

from app.api.deps import require_roles
from app.api.v1.endpoints.citas import _construir_agenda_extendida, _parse_fecha_param
from app.core.datetime_utils import formatear_fecha_emision, formatear_hora_12
from app.db.session import get_session
from app.models import Cita, Medico, Paciente, RolUsuario, Usuario

router = APIRouter(prefix="/reportes", tags=["reportes"])

_staff = require_roles(RolUsuario.secretaria, RolUsuario.admin, RolUsuario.medico)
_secretaria = require_roles(RolUsuario.secretaria, RolUsuario.admin)

_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8" />
<title>Reporte de Citas — SGCM</title>
<style>
  @page { size: A4; margin: 1.8cm; }
  body { font-family: 'Helvetica', sans-serif; color: #1f2937; font-size: 11pt; }
  h1 { color: #1e40af; margin-bottom: 0; font-size: 18pt; }
  .sub { color: #6b7280; margin-top: 4px; font-size: 10pt; }
  .emision { color: #9ca3af; font-size: 9pt; margin-top: 2px; }
  table { width: 100%; border-collapse: collapse; margin-top: 18px; }
  th { background: #1e40af; color: white; text-align: left; padding: 8px; font-size: 10pt; }
  td { padding: 7px 8px; border-bottom: 1px solid #e5e7eb; font-size: 10pt; }
  tr:nth-child(even) td { background: #f9fafb; }
  .estado { padding: 2px 8px; border-radius: 10px; font-size: 9pt; font-weight: bold; }
  .pendiente { background: #dbeafe; color: #1e3a8a; }
  .atendida { background: #dcfce7; color: #166534; }
  .cancelada { background: #f3f4f6; color: #4b5563; }
  .resumen-periodo { margin-top: 40px; background: #f9fafb;
                     padding: 14px 18px; border-radius: 6px;
                     border: 1px solid #e5e7eb; }
  .resumen-periodo h2 { color: #1e40af; font-size: 12pt; margin: 0 0 10px 0; }
  .resumen-table { width: auto; min-width: 260px; margin: 0; }
  .resumen-table td { background: transparent !important; border-bottom: none;
                      padding: 4px 14px 4px 0; }
  .resumen-table td:nth-child(2) { text-align: right; font-variant-numeric: tabular-nums; }
  .resumen-table tr.total td { font-weight: bold; border-top: 1px solid #d1d5db;
                               padding-top: 8px; }
  footer { position: fixed; bottom: 0; left: 0; right: 0; text-align: center;
           color: #9ca3af; font-size: 8pt; }
</style>
</head>
<body>
  <h1>Reporte de Citas Médicas</h1>
  <p class="emision">Reporte generado el {{ fecha_emision }}</p>
  <div class="sub">
    Hospital Regional Traumatológico y Quirúrgico Prof. Juan Bosch — SGCM<br/>
    Rango: {{ desde }} a {{ hasta }}
    {% if medico_nombre %}· Médico: {{ medico_nombre }}{% endif %}<br/>
    Total: {{ filas|length }} cita(s)
  </div>

  <table>
    <thead>
      <tr>
        <th>#</th><th>Fecha</th><th>Hora</th><th>Paciente</th><th>Médico</th><th>Estado</th>
      </tr>
    </thead>
    <tbody>
      {% for r in filas %}
      <tr>
        <td>{{ loop.index }}</td>
        <td>{{ r.fecha }}</td>
        <td>{{ r.hora }}</td>
        <td>{{ r.paciente }}</td>
        <td>{{ r.medico }}</td>
        <td><span class="estado {{ r.estado }}">{{ r.estado }}</span></td>
      </tr>
      {% endfor %}
    </tbody>
  </table>

  <div class="resumen-periodo">
    <h2>Resumen del periodo</h2>
    <table class="resumen-table">
      <tr><td>Citas pendientes</td><td>{{ resumen.pendientes }}</td></tr>
      <tr><td>Citas atendidas</td><td>{{ resumen.atendidas }}</td></tr>
      <tr><td>Citas canceladas</td><td>{{ resumen.canceladas }}</td></tr>
      <tr class="total"><td>Total general</td><td>{{ resumen.total }}</td></tr>
    </table>
  </div>

  <footer>SGCM — Generado automáticamente · HTQPJB · La Vega, R.D.</footer>
</body>
</html>
"""


@router.get("/citas.pdf")
def reporte_citas_pdf(
    desde: date_type = Query(...),
    hasta: date_type = Query(...),
    id_medico: int | None = None,
    session: Session = Depends(get_session),
    _: Usuario = Depends(_staff),
):
    stmt = select(Cita, Paciente, Medico).where(
        Cita.id_paciente == Paciente.id,
        Cita.id_medico == Medico.id,
        Cita.fecha >= desde,
        Cita.fecha <= hasta,
    )
    if id_medico:
        stmt = stmt.where(Cita.id_medico == id_medico)
    rows = session.exec(stmt.order_by(Cita.fecha, Cita.hora)).all()

    filas = [
        {
            "id": c.id,
            "fecha": c.fecha.isoformat(),
            "hora": formatear_hora_12(c.hora),
            "paciente": f"{p.nombre} {p.apellidos}",
            "medico": m.nombre,
            "estado": c.estado.value,
        }
        for c, p, m in rows
    ]

    medico_nombre = None
    if id_medico:
        m = session.get(Medico, id_medico)
        medico_nombre = m.nombre if m else None

    fecha_emision = formatear_fecha_emision()

    resumen = {
        "pendientes": sum(1 for f in filas if f["estado"] == "pendiente"),
        "atendidas":  sum(1 for f in filas if f["estado"] == "atendida"),
        "canceladas": sum(1 for f in filas if f["estado"] == "cancelada"),
        "total":      len(filas),
    }

    html_str = Template(_TEMPLATE).render(
        desde=desde.isoformat(), hasta=hasta.isoformat(),
        filas=filas, medico_nombre=medico_nombre,
        fecha_emision=fecha_emision,
        resumen=resumen,
    )
    pdf_bytes = HTML(string=html_str).write_pdf()

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="reporte_citas_{desde}_{hasta}.pdf"'},
    )


# ═════════════════════════════════════════════════════════════════
# Reportes de Agenda Extendida (secretaria/admin) — PDF y Excel
# ═════════════════════════════════════════════════════════════════

_AGENDA_TEMPLATE = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8" />
<title>Agenda del Día — SGCM</title>
<style>
  @page { size: A4 landscape; margin: 1.4cm; }
  body { font-family: 'Helvetica', sans-serif; color: #1f2937; font-size: 10pt; }
  h1 { color: #1e40af; margin: 0 0 4px 0; font-size: 17pt; }
  .institucion { color: #4b5563; font-size: 9.5pt; margin: 0; }
  .emision { color: #9ca3af; font-size: 8.5pt; margin: 2px 0 0 0; }
  .filtros { background: #f3f4f6; padding: 10px 14px; border-radius: 6px;
             margin-top: 14px; border: 1px solid #e5e7eb; font-size: 9.5pt; }
  .filtros strong { color: #1e40af; }
  .resumen-strip { display: table; width: 100%; margin: 12px 0; border-collapse: collapse; }
  .resumen-strip .cell { display: table-cell; text-align: center;
                         padding: 8px 6px; border-right: 1px solid #e5e7eb;
                         background: #f9fafb; }
  .resumen-strip .cell:last-child { border-right: none; }
  .resumen-strip .label { font-size: 8.5pt; color: #6b7280; text-transform: uppercase; letter-spacing: 0.04em; }
  .resumen-strip .value { font-size: 16pt; font-weight: bold; color: #1f2937;
                          font-variant-numeric: tabular-nums; }
  .resumen-strip .pendientes .value { color: #1e40af; }
  .resumen-strip .atendidas  .value { color: #166534; }
  .resumen-strip .canceladas .value { color: #6b7280; }
  table { width: 100%; border-collapse: collapse; margin-top: 8px; }
  th { background: #1e40af; color: white; text-align: left; padding: 7px 8px; font-size: 9pt; }
  td { padding: 6px 8px; border-bottom: 1px solid #e5e7eb; font-size: 9pt; }
  tr:nth-child(even) td { background: #f9fafb; }
  .estado { padding: 2px 8px; border-radius: 10px; font-size: 8.5pt; font-weight: bold; }
  .pendiente { background: #dbeafe; color: #1e3a8a; }
  .atendida  { background: #dcfce7; color: #166534; }
  .cancelada { background: #f3f4f6; color: #4b5563; }
  .pie-totales { margin-top: 20px; padding-top: 10px; border-top: 2px solid #1e40af;
                 text-align: right; font-size: 10pt; }
  .pie-totales strong { color: #1e40af; }
  footer { position: fixed; bottom: 0; left: 0; right: 0; text-align: center;
           color: #9ca3af; font-size: 8pt; }
</style>
</head>
<body>
  <h1>Agenda del Día</h1>
  <p class="institucion">
    Hospital Regional Traumatológico y Quirúrgico Prof. Juan Bosch — SGCM<br/>
    La Vega, República Dominicana
  </p>
  <p class="emision">Reporte generado el {{ fecha_emision }}</p>

  <div class="filtros">
    <strong>Filtros aplicados:</strong>
    Médico: {{ f_medico }} ·
    Especialidad: {{ f_especialidad }} ·
    Estado: {{ f_estado }} ·
    Rango: {{ f_desde }} a {{ f_hasta }}
  </div>

  <div class="resumen-strip">
    <div class="cell"><div class="label">Total</div><div class="value">{{ resumen.total }}</div></div>
    <div class="cell pendientes"><div class="label">Pendientes</div><div class="value">{{ resumen.pendientes }}</div></div>
    <div class="cell atendidas"><div class="label">Atendidas</div><div class="value">{{ resumen.atendidas }}</div></div>
    <div class="cell canceladas"><div class="label">Canceladas</div><div class="value">{{ resumen.canceladas }}</div></div>
  </div>

  {% if citas %}
  <table>
    <thead>
      <tr>
        <th>#</th>
        <th>Fecha</th>
        <th>Hora</th>
        <th>Paciente</th>
        <th>Médico</th>
        <th>Especialidad</th>
        <th>Estado</th>
      </tr>
    </thead>
    <tbody>
      {% for c in citas %}
      <tr>
        <td>{{ loop.index }}</td>
        <td>{{ c.fecha }}</td>
        <td>{{ c.hora_12h }}</td>
        <td>{{ c.paciente_nombre }}</td>
        <td>Dr. {{ c.medico_nombre }}</td>
        <td>{{ c.medico_especialidad }}</td>
        <td><span class="estado {{ c.estado.value }}">{{ c.estado.value }}</span></td>
      </tr>
      {% endfor %}
    </tbody>
  </table>
  {% else %}
  <p style="color:#9ca3af; margin-top:18px; text-align:center;">Sin citas para los filtros aplicados.</p>
  {% endif %}

  <div class="pie-totales">
    Total general: <strong>{{ resumen.total }}</strong> cita(s) ·
    {{ resumen.pendientes }} pendiente(s) ·
    {{ resumen.atendidas }} atendida(s) ·
    {{ resumen.canceladas }} cancelada(s)
  </div>

  <footer>SGCM — Generado automáticamente · HTQPJB · La Vega, R.D.</footer>
</body>
</html>
"""


def _filtro_medico_legible(session: Session, id_medico: int | None, busqueda: str | None) -> str:
    if id_medico:
        m = session.get(Medico, id_medico)
        return f"Dr. {m.nombre}" if m else f"#{id_medico}"
    if busqueda:
        return f"contiene «{busqueda}»"
    return "Todos"


@router.get("/agenda/pdf")
def reporte_agenda_pdf(
    id_medico: int | None = None,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    estado: str | None = Query(None),
    especialidad: str | None = None,
    busqueda_medico: str | None = None,
    session: Session = Depends(get_session),
    _: Usuario = Depends(_secretaria),
):
    """PDF de agenda extendida con los mismos filtros que /citas/agenda-extendida."""
    fd = _parse_fecha_param(fecha_desde)
    fh = _parse_fecha_param(fecha_hasta)
    agenda = _construir_agenda_extendida(
        session,
        id_medico=id_medico,
        fecha_desde=fd,
        fecha_hasta=fh,
        estado=estado,
        especialidad=especialidad,
        busqueda_medico=busqueda_medico,
    )

    html_str = Template(_AGENDA_TEMPLATE).render(
        fecha_emision=formatear_fecha_emision(),
        f_medico=_filtro_medico_legible(session, id_medico, busqueda_medico),
        f_especialidad=especialidad or "Todas",
        f_estado=(estado or "todos").capitalize(),
        f_desde=fd.isoformat() if fd else "—",
        f_hasta=fh.isoformat() if fh else "—",
        citas=agenda.citas,
        resumen={
            "total": agenda.total,
            "pendientes": agenda.pendientes,
            "atendidas": agenda.atendidas,
            "canceladas": agenda.canceladas,
        },
    )
    pdf_bytes = HTML(string=html_str).write_pdf()

    nombre = f"agenda_{fd or 'inicio'}_{fh or 'fin'}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{nombre}"'},
    )


def _generar_excel_agenda(
    *,
    citas: list,
    resumen: dict,
    filtros: dict,
    fecha_emision: str,
) -> bytes:
    """Construye el .xlsx con la misma estructura que el PDF."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Agenda"

    azul = PatternFill("solid", fgColor="1E40AF")
    blanco_neg = Font(color="FFFFFF", bold=True, size=11)
    bold = Font(bold=True)

    # Cabecera institucional
    ws["A1"] = "Agenda del Día — SGCM"
    ws["A1"].font = Font(bold=True, size=16, color="1E40AF")
    ws.merge_cells("A1:G1")
    ws["A2"] = "Hospital Regional Traumatológico y Quirúrgico Prof. Juan Bosch · La Vega, R.D."
    ws.merge_cells("A2:G2")
    ws["A3"] = f"Reporte generado el {fecha_emision}"
    ws["A3"].font = Font(italic=True, color="9CA3AF", size=10)
    ws.merge_cells("A3:G3")

    # Filtros aplicados
    ws["A5"] = "Filtros aplicados:"
    ws["A5"].font = bold
    ws["A6"] = f"Médico: {filtros['medico']}"
    ws["A7"] = f"Especialidad: {filtros['especialidad']}"
    ws["A8"] = f"Estado: {filtros['estado']}"
    ws["A9"] = f"Rango: {filtros['desde']} a {filtros['hasta']}"

    # Resumen
    ws["A11"] = "Resumen"
    ws["A11"].font = bold
    encabezados_resumen = ["Total", "Pendientes", "Atendidas", "Canceladas"]
    valores_resumen = [resumen["total"], resumen["pendientes"], resumen["atendidas"], resumen["canceladas"]]
    for i, (label, val) in enumerate(zip(encabezados_resumen, valores_resumen)):
        c_lbl = ws.cell(row=12, column=i + 1, value=label)
        c_val = ws.cell(row=13, column=i + 1, value=val)
        c_lbl.font = Font(bold=True, color="FFFFFF")
        c_lbl.fill = azul
        c_lbl.alignment = Alignment(horizontal="center")
        c_val.alignment = Alignment(horizontal="center")
        c_val.font = Font(bold=True, size=14)

    # Tabla
    columnas = ["#", "Fecha", "Hora", "Paciente", "Médico", "Especialidad", "Estado"]
    fila_inicio = 16
    for i, col in enumerate(columnas, start=1):
        c = ws.cell(row=fila_inicio, column=i, value=col)
        c.font = blanco_neg
        c.fill = azul
        c.alignment = Alignment(horizontal="left")

    for idx, c in enumerate(citas, start=1):
        fila = fila_inicio + idx
        ws.cell(row=fila, column=1, value=idx)
        ws.cell(row=fila, column=2, value=c.fecha.isoformat())
        ws.cell(row=fila, column=3, value=c.hora_12h)
        ws.cell(row=fila, column=4, value=c.paciente_nombre)
        ws.cell(row=fila, column=5, value=f"Dr. {c.medico_nombre}")
        ws.cell(row=fila, column=6, value=c.medico_especialidad)
        ws.cell(row=fila, column=7, value=c.estado.value)

    # Pie
    fila_pie = fila_inicio + len(citas) + 2
    ws.cell(row=fila_pie, column=1, value="Total general:").font = bold
    ws.cell(row=fila_pie, column=2, value=resumen["total"]).font = bold

    # Anchos
    anchos = [5, 12, 12, 28, 28, 28, 12]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/agenda/excel")
def reporte_agenda_excel(
    id_medico: int | None = None,
    fecha_desde: str | None = Query(None),
    fecha_hasta: str | None = Query(None),
    estado: str | None = Query(None),
    especialidad: str | None = None,
    busqueda_medico: str | None = None,
    session: Session = Depends(get_session),
    _: Usuario = Depends(_secretaria),
):
    """Excel (.xlsx) de agenda extendida — mismos filtros que el PDF."""
    fd = _parse_fecha_param(fecha_desde)
    fh = _parse_fecha_param(fecha_hasta)
    agenda = _construir_agenda_extendida(
        session,
        id_medico=id_medico,
        fecha_desde=fd,
        fecha_hasta=fh,
        estado=estado,
        especialidad=especialidad,
        busqueda_medico=busqueda_medico,
    )

    xlsx_bytes = _generar_excel_agenda(
        citas=agenda.citas,
        resumen={
            "total": agenda.total,
            "pendientes": agenda.pendientes,
            "atendidas": agenda.atendidas,
            "canceladas": agenda.canceladas,
        },
        filtros={
            "medico": _filtro_medico_legible(session, id_medico, busqueda_medico),
            "especialidad": especialidad or "Todas",
            "estado": (estado or "todos").capitalize(),
            "desde": fd.isoformat() if fd else "—",
            "hasta": fh.isoformat() if fh else "—",
        },
        fecha_emision=formatear_fecha_emision(),
    )

    nombre = f"agenda_{fd or 'inicio'}_{fh or 'fin'}.xlsx"
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )
